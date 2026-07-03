from __future__ import annotations

import argparse
import copy
import json
import os
import re
import shutil
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


STATUS_VALUE = "完了"
SHEET_DATE_CELL = "B5"
DATA_START_ROW = 8
SEPARATOR = "、"
COMMIT_MARKER = "__COMMIT__"
BODY_MARKER = "__BODY__"
FILES_MARKER = "__FILES__"
HEADER_SEPARATOR = "\x1f"

SKIP_SUBJECTS = (
    re.compile(r"^ok$", re.IGNORECASE),
    re.compile(r"^merge pull request", re.IGNORECASE),
    re.compile(r"^revert\b", re.IGNORECASE),
)

AREA_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"materialprinting|material printing|material print", re.IGNORECASE), "MaterialPrinting画面"),
    (re.compile(r"productinspectiongroupededitor|product inspection grouped editor", re.IGNORECASE), "ProductInspectionGroupedEditor"),
    (re.compile(r"productinspectiontable|product inspection table", re.IGNORECASE), "ProductInspectionTable"),
    (re.compile(r"dailyreportsprintfiscal", re.IGNORECASE), "DailyReportsPrintFiscal"),
    (re.compile(r"product inspection print|inspection print", re.IGNORECASE), "inspection印刷"),
    (re.compile(r"admin?tensiletests|tensiletests|tensile sidebar|tensile", re.IGNORECASE), "tensile画面"),
    (re.compile(r"pdfdocumentservice|pdfdocumentscontroller|audit zip|audit package zip", re.IGNORECASE), "PDF/ZIP出力"),
    (re.compile(r"pdf-viewer|pdf viewer|dbheadpdfviewerhost", re.IGNORECASE), "PDF viewer"),
    (re.compile(r"dbhbrowserhelpers|index\.html|dbheadpdfbridge", re.IGNORECASE), "共通スクリプト"),
    (re.compile(r"app\.css", re.IGNORECASE), "共通スタイル"),
    (re.compile(r"dailyreportsentry", re.IGNORECASE), "DailyReportsEntry"),
)

OBJECT_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"dailyreportsprintfiscal", re.IGNORECASE), "DailyReportsPrintFiscal"),
    (re.compile(r"productinspectionsheetgrouper", re.IGNORECASE), "inspection印刷のロジック"),
    (re.compile(r"production month picker", re.IGNORECASE), "production month picker component"),
    (re.compile(r"unsaved changes dialog", re.IGNORECASE), "未保存変更ダイアログ"),
    (re.compile(r"inspection toolbar spacing", re.IGNORECASE), "inspection toolbar の余白"),
    (re.compile(r"inspection toolbar", re.IGNORECASE), "inspection toolbar"),
    (re.compile(r"material printing workspace", re.IGNORECASE), "MaterialPrintingワークスペース"),
    (re.compile(r"material printing studio layout", re.IGNORECASE), "MaterialPrinting studio レイアウト"),
    (re.compile(r"audit package zip download", re.IGNORECASE), "audit ZIPダウンロード"),
    (re.compile(r"audit zip browser download", re.IGNORECASE), "audit ZIPのブラウザダウンロード"),
    (re.compile(r"audit zip validation diagnostics", re.IGNORECASE), "audit ZIP診断"),
    (re.compile(r"audit zip internal paths", re.IGNORECASE), "audit ZIP内部パス"),
    (re.compile(r"audit zip writer compile error", re.IGNORECASE), "audit ZIP writer"),
    (re.compile(r"windows explorer", re.IGNORECASE), "Windows Explorer"),
    (re.compile(r"productinspectiongroupededitor|product inspection grouped editor", re.IGNORECASE), "ProductInspectionGroupedEditor"),
    (re.compile(r"tensile sidebar", re.IGNORECASE), "tensile sidebar"),
    (re.compile(r"all companies option", re.IGNORECASE), "全会社オプション"),
    (re.compile(r"hidden material printing ui", re.IGNORECASE), "非表示MaterialPrinting UI"),
    (re.compile(r"dossier labels", re.IGNORECASE), "dossier ラベル"),
    (re.compile(r"dossier categories", re.IGNORECASE), "dossier 区分"),
    (re.compile(r"pdf export helper", re.IGNORECASE), "PDF export helper"),
    (re.compile(r"auth browser helper", re.IGNORECASE), "auth browser helper"),
    (re.compile(r"pdf actions and rename flow", re.IGNORECASE), "PDF操作と名称変更フロー"),
    (re.compile(r"load in productinspectiongroupededitor", re.IGNORECASE), "ProductInspectionGroupedEditor読込"),
)

ACTION_PATTERNS: tuple[tuple[re.Pattern[str], str, str], ...] = (
    (re.compile(r"\b(add|create)\b", re.IGNORECASE), "追加する", "追加"),
    (re.compile(r"\b(remove|delete)\b", re.IGNORECASE), "削除する", "削除"),
    (re.compile(r"\b(separate|isolate|split)\b", re.IGNORECASE), "分離する", "分離"),
    (re.compile(r"\b(share|common)\b", re.IGNORECASE), "共通化する", "共通化"),
    (re.compile(r"\b(move)\b", re.IGNORECASE), "移動する", "移動"),
    (re.compile(r"\b(rename)\b", re.IGNORECASE), "名称変更する", "名称変更"),
    (re.compile(r"\b(normalize|unify|standardize)\b", re.IGNORECASE), "統一する", "統一"),
    (re.compile(r"\b(regenerate)\b", re.IGNORECASE), "再生成する", "再生成"),
    (re.compile(r"\b(redesign|restyle)\b", re.IGNORECASE), "再設計する", "再設計"),
    (re.compile(r"\b(adjust|tune)\b", re.IGNORECASE), "調整する", "調整"),
    (re.compile(r"\b(fix)\b", re.IGNORECASE), "修正する", "修正"),
    (re.compile(r"\b(harden|prevent|improve)\b", re.IGNORECASE), "改善する", "改善"),
    (re.compile(r"\b(update|match|keep|load|open|render|embed)\b", re.IGNORECASE), "更新する", "更新"),
)

SUBJECT_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"enhance fiscal transform preview styles and structure in (.+)", re.IGNORECASE), "DailyReportsPrintFiscalの年度変換プレビューを調整"),
    (re.compile(r"refactor layout and styles for production month picker components", re.IGNORECASE), "production month picker component のレイアウトを調整"),
    (re.compile(r"update unsaved changes dialog button styles and reorder report type options", re.IGNORECASE), "未保存変更ダイアログと帳票種別の表示順を調整"),
    (re.compile(r"simplify success messages across multiple pages to a consistent format", re.IGNORECASE), "複数画面の成功メッセージ表示を統一"),
    (re.compile(r"implement unsaved changes prompt dialog and state management across multiple pages", re.IGNORECASE), "複数画面に未保存変更確認を追加"),
    (re.compile(r"improve lot grouping logic in (.+)", re.IGNORECASE), "inspection印刷のロットグループ処理を改善"),
    (re.compile(r"enhance product inspection features with dbringshipmentbox integration", re.IGNORECASE), "ProductInspectionにDbRingShipmentBox連携を追加"),
)

BODY_LINE_PATTERNS = (
    (re.compile(r"^(add|create)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を追加する"),
    (re.compile(r"^(update)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を更新する"),
    (re.compile(r"^(fix)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を修正する"),
    (re.compile(r"^(remove|delete)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を削除する"),
    (re.compile(r"^(rename)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}の名前を変更する"),
    (re.compile(r"^(move)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を移動する"),
    (re.compile(r"^(refactor)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}をリファクタリングする"),
    (re.compile(r"^(improve|enhance)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を改善する"),
    (re.compile(r"^(adjust|tune)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を調整する"),
    (re.compile(r"^(simplify)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を簡単にする"),
    (re.compile(r"^(normalize|standardize|unify)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を統一する"),
    (re.compile(r"^(implement)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を実装する"),
    (re.compile(r"^(prevent|avoid)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を防ぐ"),
    (re.compile(r"^(allow|support)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}に対応する"),
    (re.compile(r"^(use)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を使う"),
    (re.compile(r"^(keep)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を維持する"),
    (re.compile(r"^(show)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を表示する"),
    (re.compile(r"^(hide)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を非表示にする"),
    (re.compile(r"^(load)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を読み込む"),
    (re.compile(r"^(save)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を保存する"),
    (re.compile(r"^(open)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を開く"),
    (re.compile(r"^(close)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を閉じる"),
    (re.compile(r"^(render)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を表示する"),
    (re.compile(r"^(return)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を返す"),
    (re.compile(r"^(replace)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}を置き換える"),
    (re.compile(r"^(convert)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}に変換する"),
    (re.compile(r"^(skip)\s+(.+)$", re.IGNORECASE), lambda m: f"{translate_body_remainder(m.group(2))}をスキップする"),
)


@dataclass(frozen=True)
class CommitInfo:
    commit_date: date
    subject: str
    body_lines: tuple[str, ...]
    files: tuple[str, ...]


@dataclass(frozen=True)
class DailyReportEntry:
    report_date: date
    status: str
    task: str
    detail: str
    memo: str
    commit_count: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill the weekly DBH report workbook from git commits.")
    parser.add_argument("--repo-path", required=True)
    parser.add_argument("--excel-path", required=True)
    parser.add_argument("--reference-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--author")
    parser.add_argument("--author-email")
    parser.add_argument("--preview-json")
    return parser.parse_args()


def run_git(repo_path: Path, *args: str) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=repo_path,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return completed.stdout.strip()


def resolve_week_window(reference_date: date) -> tuple[date, date]:
    diff = reference_date.weekday()
    monday = reference_date - timedelta(days=diff)
    friday = monday + timedelta(days=4)
    if reference_date.weekday() >= 5:
        week_end = friday
    else:
        week_end = min(reference_date, friday)
    return monday, week_end


def should_skip_subject(subject: str) -> bool:
    text = subject.strip()
    return any(pattern.search(text) for pattern in SKIP_SUBJECTS)


def collect_commits(
    repo_path: Path,
    start_date: date,
    end_date: date,
    author_name: str | None,
    author_email: str | None,
) -> tuple[str | None, str | None, list[CommitInfo]]:
    resolved_author_name = author_name.strip() if author_name and author_name.strip() else None
    resolved_author_email = author_email.strip() if author_email and author_email.strip() else None

    if not resolved_author_name:
        resolved_author_name = run_git(repo_path, "config", "user.name") or None
    if not resolved_author_email:
        resolved_author_email = run_git(repo_path, "config", "user.email") or None

    args = [
        "log",
        f"--since={start_date:%Y-%m-%d} 00:00:00",
        f"--until={end_date:%Y-%m-%d} 23:59:59",
        "--date=short",
        "--name-only",
        f"--pretty=format:{COMMIT_MARKER}%n%ad%x1f%an%x1f%ae%x1f%s%n{BODY_MARKER}%n%b%n{FILES_MARKER}",
    ]
    if resolved_author_name:
        args.append(f"--author={resolved_author_name}")

    output = run_git(repo_path, *args)
    commits = parse_commits(output, resolved_author_email)
    return resolved_author_name, resolved_author_email, commits


def parse_commits(output: str, author_email: str | None) -> list[CommitInfo]:
    commits: list[CommitInfo] = []
    header: list[str] | None = None
    current_body_lines: list[str] = []
    current_files: list[str] = []
    in_body = False
    in_files = False

    for raw_line in output.splitlines():
        line = raw_line.rstrip("\n")
        if line == COMMIT_MARKER:
            try_add_commit(commits, header, current_body_lines, current_files, author_email)
            header = None
            current_body_lines = []
            current_files = []
            in_body = False
            in_files = False
            continue

        if header is None:
            if not line.strip():
                continue
            header = line.split(HEADER_SEPARATOR, 3)
            continue

        if line == BODY_MARKER:
            in_body = True
            in_files = False
            continue

        if line == FILES_MARKER:
            in_body = False
            in_files = True
            continue

        if in_body:
            current_body_lines.append(line.rstrip())
            continue

        if in_files and line.strip():
            current_files.append(line.strip())

    try_add_commit(commits, header, current_body_lines, current_files, author_email)
    return commits


def try_add_commit(
    commits: list[CommitInfo],
    header: list[str] | None,
    body_lines: list[str],
    files: list[str],
    author_email: str | None,
) -> None:
    if not header or len(header) < 4:
        return

    raw_date, _author_name, commit_email, subject = header
    subject = subject.strip()

    if author_email and commit_email.strip().lower() != author_email.strip().lower():
        return
    if should_skip_subject(subject):
        return

    commits.append(
        CommitInfo(
            commit_date=datetime.strptime(raw_date, "%Y-%m-%d").date(),
            subject=subject,
            body_lines=tuple(body_lines),
            files=tuple(files),
        )
    )


def resolve_sheet_title(target_date: date) -> str:
    return target_date.replace(day=1).strftime("%d-%m-%Y")


def load_preview_reports(preview_json_path: Path) -> list[DailyReportEntry]:
    payload = json.loads(preview_json_path.read_text(encoding="utf-8-sig"))
    reports: list[DailyReportEntry] = []
    for item in payload:
        reports.append(
            DailyReportEntry(
                report_date=datetime.strptime(item["ReportDate"], "%Y-%m-%d").date(),
                status=item["Status"],
                task=item["Task"],
                detail=item["Detail"],
                memo=item["Memo"],
                commit_count=int(item["CommitCount"]),
            )
        )
    return reports


def compose_daily_reports(start_date: date, end_date: date, commits: Sequence[CommitInfo]) -> list[DailyReportEntry]:
    by_day: dict[date, list[CommitInfo]] = {}
    for commit in commits:
        by_day.setdefault(commit.commit_date, []).append(commit)

    reports: list[DailyReportEntry] = []
    current = start_date
    while current <= end_date:
        day_commits = by_day.get(current, [])
        if day_commits:
            reports.append(build_daily_report(current, day_commits))
        current += timedelta(days=1)
    return reports


def build_daily_report(report_date: date, commits: Sequence[CommitInfo]) -> DailyReportEntry:
    area_counter = Counter(detect_area((commit.subject, *commit.files)) for commit in commits)
    dominant_area = area_counter.most_common(1)[0][0] if area_counter else "共通対応"

    action_counter = Counter(detect_action(commit.subject)[0] for commit in commits)
    dominant_action = "更新する" if len(action_counter) > 1 else action_counter.most_common(1)[0][0]

    translated: list[str] = []
    for commit in commits:
        text = build_memo_text(commit)
        if text not in translated:
            translated.append(text)
        if len(translated) == 6:
            break

    return DailyReportEntry(
        report_date=report_date,
        status=STATUS_VALUE,
        task=f"{dominant_area}を{dominant_action}",
        detail=build_detail(commits),
        memo=SEPARATOR.join(translated),
        commit_count=len(commits),
    )


def build_detail(commits: Sequence[CommitInfo]) -> str:
    details: list[str] = []
    seen: set[str] = set()
    for commit in commits:
        label = detect_object(commit.subject, commit.files)
        if label in seen:
            continue
        seen.add(label)
        details.append(label)
        if len(details) == 3:
            break
    if details:
        return SEPARATOR.join(details)

    areas = [detect_area((commit.subject, *commit.files)) for commit in commits]
    unique_areas = list(dict.fromkeys(areas))
    return SEPARATOR.join(unique_areas[:2])


def detect_area(texts: Iterable[str]) -> str:
    combined = " ".join(texts)
    for pattern, label in AREA_PATTERNS:
        if pattern.search(combined):
            return label
    return "共通対応"


def detect_action(subject: str) -> tuple[str, str]:
    for pattern, verb, noun in ACTION_PATTERNS:
        if pattern.search(subject):
            return verb, noun
    return "更新する", "更新"


def detect_object(subject: str, files: Sequence[str]) -> str:
    for pattern, label in OBJECT_PATTERNS:
        if pattern.search(subject):
            return label
    return detect_area((subject, *files))


def translate_subject(subject: str, files: Sequence[str]) -> str:
    trimmed = subject.strip()
    for pattern, translated in SUBJECT_PATTERNS:
        if pattern.search(trimmed):
            return translated

    verb, noun = detect_action(trimmed)
    target = detect_object(trimmed, files)

    if noun == "移動" and target.endswith("移動"):
        return f"{target}を対応"
    if noun == "再設計":
        return f"{target}を再設計"
    if noun == "再生成":
        return f"{target}を再生成"
    if noun == "共通化":
        return f"{target}を共通化"
    if noun == "統一":
        return f"{target}を統一"
    if noun == "削除":
        return f"{target}を削除"
    if noun == "追加":
        return f"{target}を追加"
    if noun == "分離":
        return f"{target}を分離"
    if noun == "名称変更":
        return f"{target}を名称変更"
    if noun == "調整":
        return f"{target}を調整"
    if noun == "修正":
        return f"{target}を修正"
    if noun == "改善":
        return f"{target}を改善"
    return f"{target}を{verb}"


def build_memo_text(commit: CommitInfo) -> str:
    subject_text = translate_subject(commit.subject, commit.files)
    body_text = format_body(commit.body_lines)
    if not body_text:
        return subject_text
    return f"{subject_text}\n{body_text}"


def format_body(body_lines: Sequence[str]) -> str:
    start = 0
    end = len(body_lines) - 1

    while start <= end and not body_lines[start].strip():
        start += 1

    while end >= start and not body_lines[end].strip():
        end -= 1

    if start > end:
        return ""

    return "\n".join(translate_body_line(line) for line in body_lines[start : end + 1])


def translate_body_line(line: str) -> str:
    if not line.strip():
        return ""

    trimmed = line.strip()
    marker_match = re.match(r"^(?P<marker>[-*]|\d+[.)])\s+(?P<content>.+)$", trimmed)
    if not marker_match:
        return translate_body_content(trimmed)

    marker = marker_match.group("marker")
    content = marker_match.group("content").strip()
    return f"{marker} {translate_body_content(content)}"


def translate_body_content(content: str) -> str:
    trimmed = content.strip()
    for pattern, builder in BODY_LINE_PATTERNS:
        match = pattern.match(trimmed)
        if match:
            return builder(match)
    return translate_body_remainder(trimmed)


def translate_body_remainder(text: str) -> str:
    translated = text.strip()
    translated = re.sub(r"\bon load\b", "読み込み時に", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bon save\b", "保存時に", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bon mobile\b", "モバイルで", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bon desktop\b", "デスクトップで", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bin preview\b", "プレビューで", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bfor print\b", "印刷用に", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bwith\b", "と", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bwithout\b", "なしで", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bafter\b", "後で", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bbefore\b", "前に", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\s+", " ", translated).strip()
    return translated


def backup_workbook(excel_path: Path) -> Path:
    backup_path = excel_path.with_name(
        f"{excel_path.stem}.backup-{datetime.now():%Y%m%d-%H%M%S}{excel_path.suffix}"
    )
    shutil.copy2(excel_path, backup_path)
    return backup_path


def choose_template_sheet(workbook) -> Worksheet:
    for sheet in workbook.worksheets:
        if sheet.title.lower().startswith("sheet"):
            return sheet
    return workbook.worksheets[-1]


def get_month_start_from_sheet(sheet: Worksheet) -> date | None:
    value = sheet[SHEET_DATE_CELL].value
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    return None


def ensure_month_sheet(workbook, template_sheet: Worksheet, target_date: date) -> Worksheet:
    expected_title = resolve_sheet_title(target_date)
    expected_month_start = target_date.replace(day=1)

    for sheet in workbook.worksheets:
        if sheet.title.lower() == expected_title.lower():
            return sheet
        month_start = get_month_start_from_sheet(sheet)
        if month_start == expected_month_start:
            return sheet

    copied = workbook.copy_worksheet(template_sheet)
    copied.title = expected_title
    copied[SHEET_DATE_CELL] = datetime(target_date.year, target_date.month, 1)
    return copied


def resolve_completed_fill(template_sheet: Worksheet) -> PatternFill:
    for row in range(DATA_START_ROW, DATA_START_ROW + 31):
        if str(template_sheet.cell(row, 3).value or "").strip() == STATUS_VALUE:
            return copy.copy(template_sheet.cell(row, 4).fill)
    return copy.copy(template_sheet.cell(DATA_START_ROW, 4).fill)


def row_for_date(target_date: date) -> int:
    return DATA_START_ROW + target_date.day - 1


def estimate_line_count(*values: str) -> int:
    segments = [value for value in values if value and value.strip()]
    if not segments:
        return 1
    longest = max(len(segment) for segment in segments)
    explicit_breaks = max(segment.count("\n") + 1 for segment in segments)
    estimated_from_length = max(1, longest // 28 + 1)
    return max(explicit_breaks, estimated_from_length)


def write_daily_report(sheet: Worksheet, report: DailyReportEntry, completed_fill_style) -> None:
    row = row_for_date(report.report_date)

    sheet.cell(row, 3).value = report.status
    sheet.cell(row, 4).value = report.task
    sheet.cell(row, 8).value = report.detail
    sheet.cell(row, 13).value = report.memo

    for column in range(4, 8):
        cell = sheet.cell(row, column)
        cell.fill = copy.copy(completed_fill_style)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column in (4, 8, 13):
        sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")

    base_height = sheet.row_dimensions[row].height or sheet.sheet_format.defaultRowHeight or 15.0
    line_count = estimate_line_count(report.task, report.detail, report.memo)
    sheet.row_dimensions[row].height = base_height * min(3, max(1, line_count))


def ranges_overlap(start_a: int, end_a: int, start_b: int, end_b: int) -> bool:
    return max(start_a, start_b) <= min(end_a, end_b)


def merge_task_spans(sheet: Worksheet, reports: Sequence[DailyReportEntry]) -> None:
    target_rows = [row_for_date(report.report_date) for report in reports]
    if not target_rows:
        return

    min_row = min(target_rows)
    max_row = max(target_rows)

    merged_to_remove = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_col == 4 and merged_range.max_col == 7 and ranges_overlap(
            merged_range.min_row, merged_range.max_row, min_row, max_row
        ):
            merged_to_remove.append(str(merged_range))

    for merged_range in merged_to_remove:
        sheet.unmerge_cells(merged_range)

    span_start = reports[0]
    previous = reports[0]

    for current in reports[1:]:
        consecutive = current.report_date == previous.report_date + timedelta(days=1)
        same_task = current.task == previous.task
        if consecutive and same_task:
            previous = current
            continue

        merge_span(sheet, span_start, previous)
        span_start = current
        previous = current

    merge_span(sheet, span_start, previous)


def merge_span(sheet: Worksheet, start: DailyReportEntry, end: DailyReportEntry) -> None:
    start_row = row_for_date(start.report_date)
    end_row = row_for_date(end.report_date)
    sheet.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=7)
    sheet.cell(start_row, 4).alignment = Alignment(wrap_text=True, vertical="top")


def write_workbook(excel_path: Path, reports: Sequence[DailyReportEntry]) -> tuple[Path, list[str]]:
    backup_path = backup_workbook(excel_path)
    workbook = load_workbook(excel_path)
    template_sheet = choose_template_sheet(workbook)
    completed_fill_style = resolve_completed_fill(template_sheet)

    touched_sheets: list[str] = []

    for report in reports:
        sheet = ensure_month_sheet(workbook, template_sheet, report.report_date)
        write_daily_report(sheet, report, completed_fill_style)
        if sheet.title not in touched_sheets:
            touched_sheets.append(sheet.title)

    grouped: dict[str, list[DailyReportEntry]] = {}
    for report in reports:
        grouped.setdefault(resolve_sheet_title(report.report_date), []).append(report)

    for sheet_title, month_reports in grouped.items():
        sheet = workbook[sheet_title]
        month_reports = sorted(month_reports, key=lambda item: item.report_date)
        merge_task_spans(sheet, month_reports)

    workbook.save(excel_path)
    workbook.close()
    return backup_path, touched_sheets


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    args = parse_args()

    repo_path = Path(args.repo_path).expanduser().resolve()
    excel_path = Path(args.excel_path).expanduser()
    reference_date = datetime.strptime(args.reference_date, "%Y-%m-%d").date()

    if not repo_path.exists():
        raise FileNotFoundError(f"Repository path not found: {repo_path}")
    if not (repo_path / ".git").exists():
        raise FileNotFoundError(f"Repository is not a git checkout: {repo_path}")
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")
    if excel_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx workbooks are supported: {excel_path}")

    week_start, week_end = resolve_week_window(reference_date)
    if args.preview_json:
        reports = load_preview_reports(Path(args.preview_json).expanduser().resolve())
        author_name = args.author
        author_email = args.author_email
        commits: list[CommitInfo] = []
    else:
        author_name, author_email, commits = collect_commits(
            repo_path, week_start, week_end, args.author, args.author_email
        )
        reports = compose_daily_reports(week_start, week_end, commits)
    backup_path, touched_sheets = write_workbook(excel_path, reports)

    print(f"repo={repo_path}")
    print(f"excel={excel_path}")
    print(f"week={week_start:%Y-%m-%d}..{week_end:%Y-%m-%d}")
    print(f"author={author_name or ''}")
    print(f"author_email={author_email or ''}")
    print(f"commits={len(commits)}")
    print(f"days_written={len(reports)}")
    print(f"backup={backup_path}")
    print(f"touched_sheets={','.join(touched_sheets)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise
